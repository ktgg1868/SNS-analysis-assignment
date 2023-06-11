from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
import requests
import time
import sys
import os

#웹드라이버 설정
print("웹드라이버 설정 시작")
path = ChromeDriverManager().install()
driver = webdriver.Chrome(path)
print("웹드라이버 설정 완료")

#파일이름에 들어갈 시간 설정
now = time.localtime()
s = '%04d년 %02d월 %02d일 %02d시 %02d분 %02d초' %(now.tm_year, now.tm_mon, now.tm_mday , now.tm_hour, now.tm_min, now.tm_sec)

#키워드 미리 할달(제출은 input으로 받아서 실행하기)
search_name = input("1.공고명으로 검색할 키워드는 무엇입니까?: ")
date_start = input("2.조회 시작일자 입력(예:2019/01/01): ")
date_end = input("3.조회 종료일자 입력(예:2019/03/31): ")
folder_path = input("4.파일로 저장할 폴더 이름을 입력하세요(예:c\py_temp\): ")
save_txt = f"{folder_path}{s}.txt"
save_xls = f"{folder_path}{s}.xls"

#입력받은 폴더경로가 없을 경우 생성
if not os.path.exists(folder_path):
    print(f"입력하신 폴더경로인 {folder_path} 가 존재하지 않아 경로 생성 후 크롤링 작업을 진행합니다.")
    os.makedirs(folder_path)
else:
    print(f"입력한 경로인 {folder_path} 가 존재하어 바로 크롤링 작업을 시작하겠습니다.")

#페이지 접속
url = "https://www.g2b.go.kr/index.jsp"
driver.get(url)
driver.implicitly_wait(5)

#키워드 입력 후 검색버튼 클릭
searchbar = driver.find_element(By.ID, "bidNm")
searchbar.clear()    
searchbar.send_keys(search_name)

s_date = driver.find_element(By.ID, "fromBidDt")
s_date.send_keys(date_start)

e_date = driver.find_element(By.ID, "toBidDt")
e_date.send_keys(date_end)

driver.find_element(By.CLASS_NAME, "btn_dark").click()

#검색 후 로딩 대기
driver.implicitly_wait(3)

#프레임이 분리가 되어있는 환경이라 일바적인 방법으로는 데이터에 접근불가
#프레임전환코를 사용하여 데이터가 담겨있는 프레임에 접근하기

driver.switch_to.default_content() #기본프레임으로 이동
main_frame = driver.find_element(By.ID, 'sub')
driver.switch_to.frame(main_frame) #기본프레임 -> 'sub'프레임으로 이동
content_frame = driver.find_element(By.NAME, 'main')
driver.switch_to.frame(content_frame) #'sub'프레임 -> 결과값이 있는 프레임으로 이동

#bs를 이용하여 페이지 코드를 가져와서 태그 정리

full_html = driver.page_source
soup = bs(full_html, 'html.parser')
contents_list = soup.find('tbody')
rows = contents_list.find_all("tr")

#txt파일 저장 과정
orig_stdout = sys.stdout
f = open(save_txt, 'a', encoding='UTF-8')
sys.stdout = f

#xls파일 저장 과정
wb = Workbook()
ws = wb.active

#10개의 행을 설정 후 각 열의 내용을 변수에 저장한 후 출력
rows = rows[:10]
i = int(1)

#xls파일의 첫번째 행 작성
thead = soup.find('thead')
header_row = thead.find('tr')
header_cells = header_row.find_all('th')
header_values = [cell.get_text(strip=True) for cell in header_cells]

ws.cell(row=1, column=2, value=header_values[0])
ws.cell(row=1, column=3, value=header_values[1])
ws.cell(row=1, column=4, value=header_values[2])
ws.cell(row=1, column=5, value=header_values[3])
ws.cell(row=1, column=6, value="공고URL주소")
ws.cell(row=1, column=7, value=header_values[4])
ws.cell(row=1, column=8, value=header_values[5])
ws.cell(row=1, column=9, value=header_values[6])
ws.cell(row=1, column=10, value=header_values[7])
ws.cell(row=1, column=11, value=header_values[8])
ws.cell(row=1, column=12, value=header_values[9])

#txt파일 및 xls파일 내용 채우
for row in rows:
    
    columns = row.find_all('td')

    first_column = columns[0].get_text(strip=True)
    second_column = columns[1].get_text(strip=True)
    third_column = columns[2].get_text(strip=True)
    fourth_column = columns[3].get_text(strip=True)
    fourth_column_href = columns[3]
    fifth_column = columns[4].get_text(strip=True)
    sixth_column = columns[5].get_text(strip=True)
    seventh_column = columns[6].get_text(strip=True)
    eighth_column = columns[7].get_text(strip=True)
    ninth_column = columns[8].get_text(strip=True)
    tenth_column = columns[9].get_text(strip=True)
    
    href = fourth_column_href.find('a')
    if href is not None:
        href_text = href.get_text(strip=True)
        href_url = href['href']

    print("%d번째 공고내용을 출력합니다. ~~~~~"%i)
    print("업무:", first_column)
    print("공고번호-차수:", second_column)
    print("분류:", third_column)
    print("공고명: ", fourth_column)
    print("URL 주소: ", href_url)
    print("공고기관: ", fifth_column)
    print("수요기관: ", sixth_column)
    print("계약방법: ", seventh_column)
    print("입력일시(입찰마감일시): ", eighth_column)
    print("공동수급: ", ninth_column)
    print("투찰여부: ", tenth_column)
    print("\n")
    
    i += 1
    
    ws.cell(row=i, column=1, value=i)
    ws.cell(row=i, column=2, value=first_column)
    ws.cell(row=i, column=3, value=second_column)
    ws.cell(row=i, column=4, value=third_column)
    ws.cell(row=i, column=5, value=fourth_column)
    ws.cell(row=i, column=6, value=href_url)
    ws.cell(row=i, column=7, value=fifth_column)
    ws.cell(row=i, column=8, value=sixth_column)
    ws.cell(row=i, column=9, value=seventh_column)
    ws.cell(row=i, column=10, value=eighth_column)
    ws.cell(row=i, column=11, value=ninth_column)
    ws.cell(row=i, column=12, value=tenth_column)
    
    

#txt파일 저장 및 종료
sys.stdout = orig_stdout
f.close()

#xls파일 저장
wb.save(save_xls)


print("크롤링 작업을 완료하였습니다.")